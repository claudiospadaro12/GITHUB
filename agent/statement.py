"""Lettura e analisi dello statement (storico trade) esportato da MT5.

Prende il "Report Cronistorico dei Trade" (.xlsx) e ne ricava:
 - la lista dei trade chiusi (posizioni), con netto = profitto + commissioni + swap
 - il collegamento trade -> strategia/EA (dai commenti degli ordini)
 - statistiche: netto totale, per simbolo, per giorno, per strategia, win rate,
   profit factor, aspettativa, migliori/peggiori.

Usato dal report settimanale del sabato.
"""

from __future__ import annotations

from collections import defaultdict
from dataclasses import dataclass, field

import openpyxl


@dataclass
class Trade:
    pid: str
    symbol: str
    side: str            # buy / sell
    volume: float
    open_time: str
    open_price: float
    close_time: str
    close_price: float
    profit: float
    commission: float
    swap: float
    net: float
    strategy: str = ""   # dedotta dal commento dell'ordine


@dataclass
class Stats:
    trades: list[Trade] = field(default_factory=list)
    net_total: float = 0.0
    gross_profit: float = 0.0
    gross_loss: float = 0.0
    wins: int = 0
    losses: int = 0
    by_symbol: dict = field(default_factory=dict)
    by_day: dict = field(default_factory=dict)
    by_strategy: dict = field(default_factory=dict)

    @property
    def win_rate(self) -> float:
        n = len(self.trades)
        return (self.wins / n) if n else 0.0

    @property
    def profit_factor(self) -> float:
        return (self.gross_profit / abs(self.gross_loss)) if self.gross_loss else float("inf")

    @property
    def expectancy(self) -> float:
        n = len(self.trades)
        return (self.net_total / n) if n else 0.0


def _num(v, default=0.0) -> float:
    try:
        return float(v)
    except (TypeError, ValueError):
        return default


def _order_comments(ws) -> dict:
    """Mappa id-ordine -> commento (strategia), dalla sezione 'Ordini'."""
    rows = list(ws.iter_rows(values_only=True))
    comments: dict[str, str] = {}
    in_orders = False
    header_idx = None
    for r in rows:
        c0 = str(r[0]) if r and r[0] is not None else ""
        if c0 == "Ordini":
            in_orders = True
            continue
        if in_orders:
            if str(r[1] or "") == "Ordine":  # header della tabella ordini
                header_idx = True
                continue
            if header_idx and r[1] is not None:
                oid = str(r[1])
                # il commento e' l'ultima cella non vuota della riga
                comment = ""
                for cell in reversed(r):
                    if cell not in (None, ""):
                        s = str(cell)
                        # scarta le celle "tecniche" tipo [sl ...] / [tp ...] / filled
                        if not s.startswith("[") and s not in ("filled", "canceled"):
                            comment = s
                        break
                if comment:
                    comments[oid] = comment
    return comments


def parse(path: str, since: str | None = None) -> Stats:
    """Legge lo statement. Se `since` (formato 'AAAA.MM.GG') e' dato, considera
    solo i trade aperti da quella data in poi (per limitarsi alla settimana)."""
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = list(ws.iter_rows(values_only=True))

    # trova la sezione "Posizioni"
    start = end = None
    for i, r in enumerate(rows):
        if r and str(r[1] or "") == "Posizione":
            start = i + 1
        if start and str(r[0] or "") == "Ordini":
            end = i
            break
    if start is None:
        raise ValueError("Sezione 'Posizioni' non trovata: non sembra un report storico MT5.")
    if end is None:
        end = len(rows)

    comments = _order_comments(ws)

    st = Stats()
    for r in rows[start:end]:
        if not r or r[1] is None:
            continue
        try:
            pid = str(r[1])
            t = Trade(
                pid=pid, symbol=str(r[2]), side=str(r[3]), volume=_num(r[4]),
                open_time=str(r[0]), open_price=_num(r[5]),
                close_time=str(r[8]), close_price=_num(r[9]),
                commission=_num(r[10]), swap=_num(r[11]), profit=_num(r[12]),
                net=0.0, strategy=comments.get(pid, ""),
            )
        except (IndexError, ValueError):
            continue
        # filtro settimana: salta i trade aperti prima di `since`
        if since and t.open_time[:10] < since:
            continue
        t.net = t.profit + t.commission + t.swap
        st.trades.append(t)

    # aggregati
    for t in st.trades:
        st.net_total += t.net
        if t.net > 0:
            st.gross_profit += t.net; st.wins += 1
        else:
            st.gross_loss += t.net; st.losses += 1
        day = t.open_time[:10]
        for key, bucket in ((t.symbol, st.by_symbol), (day, st.by_day),
                            (t.strategy or "(manuale/senza commento)", st.by_strategy)):
            b = bucket.setdefault(key, {"n": 0, "wins": 0, "net": 0.0})
            b["n"] += 1; b["wins"] += 1 if t.net > 0 else 0; b["net"] += t.net

    return st
